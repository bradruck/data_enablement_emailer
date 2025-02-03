# excel_manager module
# Module holds the class => ExcelManager - manages the Excel File Interface
# Class responsible for all the excel file interface, including file name search and data pull
#
from glob import glob
from openpyxl import load_workbook


class ExcelManager(object):
    def __init__(self):
        self.account_data = {}
        self.account_file_name = ""

    # Open workbook and search for row number that corresponds to the ticket key
    #
    def excel_search(self, ticket, path):
        self.account_file_name = self.get_file_name('{}/*.xlsx'.format(path))
        wb = load_workbook(filename=self.account_file_name, data_only=True)
        sheet = wb['Sheet1']
        for row in range(1, sheet.max_row+1):
            if sheet['A' + str(row)].value == ticket.key:
                return row

    # With the matched row number as input, read and save account data for email population
    #
    def excel_read(self, row):
        wb = load_workbook(filename=self.account_file_name)
        sheet = wb['Sheet1']
        self.account_data = {
            "market_id":            sheet['B' + str(row)].value,
            "beacon_id":            sheet['D' + str(row)].value,
            "data_contract_id":     sheet['G' + str(row)].value
        }
        return self.account_data

    # Search specified zfs folder for what should be the only file, get name and return
    #
    @staticmethod
    def get_file_name(path):
        for file_name in glob(path):
            return file_name
